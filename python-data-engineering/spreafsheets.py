import openpyxl
from operator import itemgetter

empl_file = openpyxl.load_workbook("employees.xlsx")
empl_list = empl_file["Sheet1"]

# remove columns 3 and 4
empl_list.delete_cols(3,4)

employees_by_experience = []

for empl_row in range(2, empl_list.max_row + 1):
    empl_name = empl_list.cell(empl_row, 1).value
    empl_experience = empl_list.cell(empl_row, 2).value

    employees_by_experience.append({ 
        "name":  empl_name, 
        "experience": int(empl_experience)
    })

# sort the list of dictionaries by experience
new_list = sorted(employees_by_experience, key=itemgetter("experience"), reverse=True)

# add entries to the spreadsheet sorted experience
for empl_row in range(2, empl_list.max_row + 1):
    empl_name = empl_list.cell(empl_row, 1)
    empl_experience = empl_list.cell(empl_row, 2)

    # because the rows start from row 2, but index for our list starts at 0
    index = empl_row - 2
    employee = new_list[index]

    empl_name.value = employee["name"]
    empl_experience.value = employee["experience"]

empl_file.save("employees_sorted_by_experience.xlsx")

